import openpyxl
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl import workbook

wb = openpyxl.load_workbook("GILDIA.xlsx")
sheet_sm = wb['SM']
sheet_lista = wb['lista_produktow']



lista_produktow = []
sklepowa_lista_produktow =[]

my_red = openpyxl.styles.colors.Color(rgb='80ff00')
my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_red)

#pobranie listy produktów bazowych
for i in range(sheet_lista.max_row):
    wartosc_produktu = sheet_lista.cell(row=(i + 1), column=1).value
    lista_produktow.append(wartosc_produktu)
#pokolorowanie wszystkich wartosci mniejszych od 6
for i in range(sheet_sm.max_row):
    if int(sheet_sm.cell(row=(i + 1), column=3).value) <= 6:
        sheet_sm.cell(row=(i + 1), column=3).fill = my_fill


#utworzenie listy produktow dla danego sklepu
for i in range(sheet_sm.max_row):
    wartosc = sheet_sm.cell(row=(i+1), column=1).value
    if wartosc == sheet_sm.cell(row=(i+2), column=1).value:
        sklepowa_lista_produktow.append(sheet_sm.cell(row=(i+1), column=2).value)

#sprawdzenie czego nie ma na liście
    if not wartosc == sheet_sm.cell(row=(i+2), column=1).value:
        czego_nie_ma = list(set(lista_produktow) - set(sklepowa_lista_produktow))
        print(czego_nie_ma)
        sklepowa_lista_produktow = []
        for ix, produkt in enumerate(czego_nie_ma):
            if (i - len(czego_nie_ma) - ix) <= 0:
                pass
            else:
                sheet_sm.cell(row=(i - len(czego_nie_ma) - ix), column=4).value = produkt

#       print(czego_nie_ma)


#wb = workbook()
wb.save(filename="gildia8.xlsx")